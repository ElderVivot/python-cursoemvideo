import openpyxl

entrada = openpyxl.load_workbook('D:\\Importador\\Conjunto de dados\\Fortes - Financeiro\\baixas.xlsx')
saida = open("D:\\Importador\\Conjunto de dados\\Fortes - Financeiro\\baixas2.txt", "w")

# guarda todas as planilhas que tem dentro do arquivo excel
planilhas = entrada.get_sheet_names()

# lê cada planilha
for p in planilhas:

    # pega o nome da planilha
    planilha = entrada.get_sheet_by_name(p)

    # pega a quantidade de linha que a planilha tem
    max_row = planilha.max_row
    # pega a quantidade de colunca que a planilha tem
    max_column = planilha.max_column

    # lê cada linha e coluna da planilha e impreme
    for i in range(1, max_row + 1):
        # lê as colunas
        for j in range(1, max_column + 1):
            # pega o valor da célula
            cell_obj = planilha.cell(row=i, column=j)
            # gera o resultado num arquivo
            resultado = str(cell_obj.value) + ';'
            resultado = resultado.replace('None', '')
            saida.write(resultado)

        # faz uma quebra de linha para passar pra nova coluna
        saida.write('\n')
# fecha o arquivo
saida.close()